#!/usr/bin/env python3
"""Render a validated hotel contract JSON into an agenda rate table.

Usage:
  python3 render.py contract.json --format xlsx --out rates.xlsx
  python3 render.py contract.json --format html --out rates.html
  python3 render.py contract.json --format csv  --out rates_daily.csv

xlsx requires openpyxl (pip install openpyxl --break-system-packages if missing).
html/csv are stdlib-only.
"""
import argparse
import csv
import html
import json
import sys
from datetime import date, timedelta


def d(s):
    y, m, dd = map(int, s.split("-"))
    return date(y, m, dd)


def xl_safe(v):
    """Neutralize spreadsheet formula injection in untrusted strings.

    A cell starting with = + - @ or tab/CR is interpreted as a formula by
    Excel/LibreOffice. Contract data is untrusted: prefix with an apostrophe.
    Also strip control and invisible characters.
    """
    if not isinstance(v, str):
        return v
    v = "".join(ch for ch in v if ch >= " " or ch == "\n")
    v = v.replace("\u200b", "").replace("\u200c", "").replace("\u200d", "") \
         .replace("\u2060", "").replace("\ufeff", "")
    if v[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + v
    return v


def build_day_index(contract):
    """Map every covered date -> season code."""
    idx = {}
    for s in contract["seasons"]:
        for r in s["ranges"]:
            cur, end = d(r["from"]), d(r["to"])
            while cur <= end:
                idx[cur] = s["code"]
                cur += timedelta(days=1)
    return idx


def collapse_periods(day_index):
    """Consecutive days with the same season -> [(from, to, season)]."""
    periods = []
    for day in sorted(day_index):
        code = day_index[day]
        if periods and periods[-1][2] == code and periods[-1][1] + timedelta(days=1) == day:
            periods[-1] = (periods[-1][0], day, code)
        else:
            periods.append((day, day, code))
    return periods


def rate_rows(contract):
    """Unique (room, board, basis, plan, occupancy) row keys, ordered."""
    seen, rows = set(), []
    for r in contract["rates"]:
        key = (r.get("segment") or "fit", r["room"], r["board"], r["basis"],
               r.get("rate_plan") or "standard",
               json.dumps(r.get("occupancy"), sort_keys=True))
        if key not in seen:
            seen.add(key)
            rows.append(key)
    rows.sort(key=lambda k: (k[0] != "fit",))  # FIT rows first, stable otherwise
    return rows


def rate_lookup(contract):
    lut = {}
    for r in contract["rates"]:
        key = (r.get("segment") or "fit", r["room"], r["board"], r["basis"],
               r.get("rate_plan") or "standard",
               json.dumps(r.get("occupancy"), sort_keys=True), r["season"])
        lut[key] = r
    return lut


def cell_text(r, currency):
    if r is None:
        return ""
    parts = [f"{r['price']:g} {currency}"]
    if r.get("min_stay"):
        parts.append(f"min {r['min_stay']}n")
    if r.get("release"):
        parts.append(f"rel {r['release']}d")
    if r.get("confidence") == "low":
        parts.append("⚠")
    return " · ".join(parts)


def occ_text(occ_json):
    occ = json.loads(occ_json) if occ_json and occ_json != "null" else None
    if not occ:
        return ""
    return f"{occ['min']}-{occ['max']} pax" if occ["min"] != occ["max"] else f"{occ['min']} pax"


def commission_amount(contract, rate):
    comm = contract.get("commission")
    if not comm or comm.get("model") != "commissionable":
        return None
    if rate.get("commissionable") is False:
        return None
    pct = comm.get("percent")
    if not isinstance(pct, (int, float)):
        return None
    return round(rate["price"] * pct / 100.0, 2)


def daily_records(contract, day_index):
    lut = rate_lookup(contract)
    rows = rate_rows(contract)
    for day in sorted(day_index):
        season = day_index[day]
        for seg, room, board, basis, plan, occ in rows:
            r = lut.get((seg, room, board, basis, plan, occ, season))
            if r:
                yield {"date": day.isoformat(), "season": season, "segment": seg,
                       "room": room,
                       "board": board, "basis": basis, "rate_plan": plan,
                       "occupancy": occ_text(occ), "price": r["price"],
                       "commission": commission_amount(contract, r),
                       "min_stay": r.get("min_stay"), "release": r.get("release"),
                       "confidence": r.get("confidence", "high")}


def render_csv(contract, day_index, out):
    fields = ["date", "season", "segment", "room", "board", "basis", "rate_plan",
              "occupancy", "price", "commission", "min_stay", "release", "confidence"]
    with open(out, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for rec in daily_records(contract, day_index):
            w.writerow({k: xl_safe(v) for k, v in rec.items()})


def policy_lines(contract):
    lines = []
    for s in contract.get("supplements") or []:
        u = "%" if s.get("unit") == "percent" else f" {contract['currency']}"
        flags = []
        if s.get("mandatory") is True:
            flags.append("mandatory")
        elif s.get("mandatory") is False:
            flags.append("optional")
        if s.get("payment"):
            flags.append(s["payment"].replace("_", " "))
        detail = f"{s.get('amount')}{u} ({s.get('basis')}) - {s.get('applies_to') or 'all'}"
        if flags:
            detail += " [" + ", ".join(flags) + "]"
        lines.append((f"Supplement/{s.get('type') or 'untyped'}", s.get("label"),
                      detail, s.get("note")))
    for s in contract.get("reductions") or []:
        u = "%" if s.get("unit") == "percent" else f" {contract['currency']}"
        lines.append(("Reduction", s.get("label"),
                      f"{s.get('amount')}{u} ({s.get('basis')}) - {s.get('applies_to') or 'all'}",
                      s.get("note")))
    for cpo in contract.get("child_policy") or []:
        val = {"free": "free",
               "percent_of_adult": f"{cpo.get('value')}% of adult",
               "fixed_price": f"{cpo.get('value')} {contract['currency']}"}.get(cpo.get("basis"), str(cpo.get("value")))
        lines.append(("Child policy", f"{cpo.get('age_from')}-{cpo.get('age_to')} yrs",
                      f"{val} ({cpo.get('condition') or ''})", cpo.get("note")))
    for cx in contract.get("cancellation") or []:
        lines.append(("Cancellation", f"≥ {cx.get('days_before')} days before",
                      f"penalty: {cx.get('penalty')}", cx.get("note")))
    for b in contract.get("blackout_dates") or []:
        lines.append(("Blackout", f"{b.get('from')} → {b.get('to')}", b.get("reason") or "", None))
    for ev in contract.get("special_events") or []:
        lines.append(("Special event", f"{ev.get('label')} ({ev.get('from')} → {ev.get('to')})",
                      ev.get("effect") or "", ev.get("note")))
    for i, promo in enumerate(contract.get("promotions") or []):
        disc = (f"-{promo.get('discount_percent')}%" if promo.get("discount_percent")
                else f"-{promo.get('discount_amount')} {contract['currency']}")
        windows = []
        if promo.get("booking_window"):
            windows.append(f"book {promo['booking_window'].get('from')} → {promo['booking_window'].get('to')}")
        if promo.get("stay_window"):
            windows.append(f"stay {promo['stay_window'].get('from')} → {promo['stay_window'].get('to')}")
        lines.append(("Promotion", promo.get("code") or promo.get("label"),
                      f"{disc} | {' | '.join(windows)} | {promo.get('conditions') or ''}"
                      + (" | not combinable" if promo.get("combinable") is False else ""),
                      promo.get("note")))
    for rst in contract.get("restrictions") or []:
        span = f"{rst.get('from')} → {rst.get('to')}" if rst.get("from") else "all periods"
        lines.append(("Restriction", rst.get("type"),
                      f"{span} | {rst.get('detail') or ''}", rst.get("note")))
    for mp in contract.get("modification_policy") or []:
        fee = ("free" if mp.get("fee") in (0, 0.0)
               else f"{mp.get('fee')} {contract['currency']}" if mp.get("fee") is not None else "n/a")
        lines.append(("Modification", f"≥ {mp.get('days_before')} days before",
                      f"{mp.get('allowed')} | fee: {fee}", mp.get("note")))
    svc = contract.get("services") or {}
    for s in svc.get("included") or []:
        lines.append(("Service included", s.get("label"), s.get("detail") or "", s.get("note")))
    for s in svc.get("excluded") or []:
        lines.append(("Service NOT included", s.get("label"), s.get("detail") or "", s.get("note")))
    pp = contract.get("pets_policy")
    if pp:
        det = []
        if pp.get("max_weight_kg"): det.append(f"max {pp['max_weight_kg']} kg")
        if pp.get("max_count"): det.append(f"max {pp['max_count']}")
        if pp.get("restrictions"): det.append(pp["restrictions"])
        if pp.get("fee_ref"): det.append(pp["fee_ref"])
        lines.append(("Pets", "allowed" if pp.get("allowed") else "not allowed",
                      " | ".join(det), pp.get("note")))
    gc = contract.get("group_conditions")
    if gc:
        if gc.get("min_pax"):
            lines.append(("Group", "Minimum pax", str(gc["min_pax"]), gc.get("note")))
        fp = gc.get("free_places")
        if fp:
            lines.append(("Group", "Free places",
                          f"1 free per {fp.get('per_paying')} paying"
                          + (f" ({fp.get('basis')})" if fp.get('basis') else ""), fp.get("note")))
        for dep in gc.get("deposit_schedule") or []:
            lines.append(("Group", f"Deposit ≥ {dep.get('days_before')} days before",
                          f"{dep.get('percent')}%", dep.get("note")))
        for cx in gc.get("cancellation") or []:
            lines.append(("Group", f"Cancellation ≥ {cx.get('days_before')} days before",
                          f"penalty: {cx.get('penalty')}", cx.get("note")))
        if gc.get("rooming_list_deadline_days"):
            lines.append(("Group", "Rooming list",
                          f"{gc['rooming_list_deadline_days']} days before arrival", None))
    return lines


def render_xlsx(contract, day_index, out):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("openpyxl not installed: pip install openpyxl --break-system-packages, "
                 "or use --format html/csv")

    periods = collapse_periods(day_index)
    rows = rate_rows(contract)
    lut = rate_lookup(contract)
    cur = contract["currency"]

    wb = Workbook()

    # --- Agenda sheet ---
    ws = wb.active
    ws.title = "Agenda"
    head_fill = PatternFill("solid", fgColor="2A3F54")
    head_font = Font(color="FFFFFF", bold=True)
    low_fill = PatternFill("solid", fgColor="FFF3CD")
    season_fills = {}
    palette = ["DCEBF7", "D4EDDA", "FCE8D5", "E8DAEF", "F9E2E2", "E2F0F9"]

    ws.cell(1, 1, f"{contract['hotel']['name']} - rates ({cur})").font = Font(bold=True, size=13)
    headers = ["Segment", "Room", "Board", "Basis", "Occ.", "Plan"] + [
        f"{f.strftime('%d/%m/%y')} → {t.strftime('%d/%m/%y')}\nSeason {sc}" for f, t, sc in periods]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(3, col, h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for ri, (seg, room, board, basis, plan, occ) in enumerate(rows, 4):
        ws.cell(ri, 1, seg.upper())
        ws.cell(ri, 2, xl_safe(room))
        ws.cell(ri, 3, board)
        ws.cell(ri, 4, basis.replace("per_", "").replace("_", " "))
        ws.cell(ri, 5, occ_text(occ))
        ws.cell(ri, 6, xl_safe(plan))
        for ci, (f, t, sc) in enumerate(periods, 7):
            r = lut.get((seg, room, board, basis, plan, occ, sc))
            cell = ws.cell(ri, ci, xl_safe(cell_text(r, cur)))
            if sc not in season_fills:
                season_fills[sc] = PatternFill("solid", fgColor=palette[len(season_fills) % len(palette)])
            if r:
                cell.fill = low_fill if r.get("confidence") == "low" else season_fills[sc]
            cell.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    for col in range(3, 7):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for col in range(7, 7 + len(periods)):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.freeze_panes = "G4"

    # --- Daily sheet ---
    ws2 = wb.create_sheet("Daily")
    fields = ["date", "season", "segment", "room", "board", "basis", "rate_plan",
              "occupancy", "price", "commission", "min_stay", "release", "confidence"]
    for col, h in enumerate(fields, 1):
        cell = ws2.cell(1, col, h)
        cell.fill = head_fill
        cell.font = head_font
    for ri, rec in enumerate(daily_records(contract, day_index), 2):
        for col, h in enumerate(fields, 1):
            ws2.cell(ri, col, xl_safe(rec[h]))
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["D"].width = 28

    # --- Contract sheet (identifiers, commission, channels, metadata) ---
    kv = []
    h = contract.get("hotel") or {}
    kv.append(("Hotel", h.get("name")))
    if h.get("location"): kv.append(("Location", h["location"]))
    if h.get("category"): kv.append(("Category", h["category"]))
    kv.append(("Currency", contract.get("currency")))
    cpd = contract.get("contract_period") or {}
    kv.append(("Contract period", f"{cpd.get('from')} → {cpd.get('to')}"))
    comm = contract.get("commission")
    if comm:
        kv.append(("Commission model", comm.get("model")))
        if comm.get("percent") is not None:
            kv.append(("Commission %", comm["percent"]))
        if comm.get("base"): kv.append(("Commission base", comm["base"]))
        if comm.get("payment_terms"): kv.append(("Commission payment", comm["payment_terms"]))
        if comm.get("note"): kv.append(("Commission note", comm["note"]))
    for ch in contract.get("booking_channels") or []:
        val = " | ".join(str(x) for x in
                         [ch.get("detail"), ch.get("identifier"), ch.get("instructions")] if x)
        kv.append((f"Channel: {ch.get('channel')}", val))
    idf = contract.get("identifiers") or {}
    if idf.get("contract_ref"): kv.append(("Contract ref", idf["contract_ref"]))
    for code_name, code_val in (idf.get("hotel_codes") or {}).items():
        if code_val:
            kv.append((f"Hotel code: {code_name}", code_val))
    for room_name, code_val in (idf.get("room_codes") or {}).items():
        kv.append((f"Room code: {room_name}", code_val))
    md = contract.get("metadata") or {}
    if md.get("keywords"): kv.append(("Keywords", ", ".join(md["keywords"])))
    dest = md.get("destination") or {}
    if dest:
        kv.append(("Destination", ", ".join(str(x) for x in
                   [dest.get("city"), dest.get("region"), dest.get("country")] if x)))
    if md.get("market"): kv.append(("Market", ", ".join(md["market"])))
    if kv:
        wsc = wb.create_sheet("Contract")
        for col, hh in enumerate(["Field", "Value"], 1):
            cell = wsc.cell(1, col, hh)
            cell.fill = head_fill
            cell.font = head_font
        for ri, (k, val) in enumerate(kv, 2):
            wsc.cell(ri, 1, xl_safe(k))
            wsc.cell(ri, 2, xl_safe(val if val is not None else ""))
        wsc.column_dimensions["A"].width = 28
        wsc.column_dimensions["B"].width = 60
        wsc.freeze_panes = "A2"

    # --- Rooms sheet (mapping catalog) ---
    catalog = contract.get("room_catalog") or []
    if catalog:
        wsr = wb.create_sheet("Rooms")
        rfields = ["raw", "code", "room_type", "category", "view", "bedding",
                   "accessible", "amenities", "capacity_hint",
                   "unmatched_tokens", "confidence", "mapped_by"]
        for col, h in enumerate(rfields, 1):
            cell = wsr.cell(1, col, h.replace("_", " "))
            cell.fill = head_fill
            cell.font = head_font
        for ri, e in enumerate(catalog, 2):
            for col, h in enumerate(rfields, 1):
                val = e.get(h)
                if isinstance(val, list):
                    val = ", ".join(val)
                elif isinstance(val, bool):
                    val = "yes" if val else "no"
                cell = wsr.cell(ri, col, xl_safe(val if val is not None else ""))
                if h == "confidence" and e.get("confidence") in ("low", "medium"):
                    cell.fill = low_fill
        wsr.freeze_panes = "A2"
        wsr.column_dimensions["A"].width = 34
        wsr.column_dimensions["B"].width = 16
        wsr.column_dimensions["H"].width = 30
        wsr.column_dimensions["J"].width = 24

    # --- Policies sheet ---
    ws3 = wb.create_sheet("Policies")
    for col, h in enumerate(["Type", "Item", "Detail", "Note"], 1):
        cell = ws3.cell(1, col, h)
        cell.fill = head_fill
        cell.font = head_font
    for ri, line in enumerate(policy_lines(contract), 2):
        for col, v in enumerate(line, 1):
            ws3.cell(ri, col, xl_safe(v or ""))
    for col, w in zip("ABCD", (16, 34, 55, 40)):
        ws3.column_dimensions[col].width = w

    wb.save(out)


def render_html(contract, day_index, out):
    periods = collapse_periods(day_index)
    rows = rate_rows(contract)
    lut = rate_lookup(contract)
    cur = contract["currency"]
    e = html.escape

    # html: segment column
    parts = ["<!doctype html><meta charset='utf-8'><style>"
             "body{font:14px system-ui;margin:24px}table{border-collapse:collapse}"
             "th,td{border:1px solid #ccc;padding:6px 10px;text-align:left}"
             "th{background:#2a3f54;color:#fff}td.low{background:#fff3cd}"
             "h2{margin-top:32px}</style>",
             f"<h1>{e(contract['hotel']['name'])} - rates ({e(cur)})</h1><table><tr>"
             "<th>Segment</th><th>Room</th><th>Board</th><th>Basis</th><th>Occ.</th><th>Plan</th>"]
    parts += [f"<th>{f.strftime('%d/%m/%y')} → {t.strftime('%d/%m/%y')}<br>Season {e(sc)}</th>"
              for f, t, sc in periods]
    parts.append("</tr>")
    for seg, room, board, basis, plan, occ in rows:
        parts.append(f"<tr><td>{e(seg.upper())}</td><td>{e(room)}</td><td>{e(board)}</td>"
                     f"<td>{e(basis.replace('per_', '').replace('_', ' '))}</td>"
                     f"<td>{e(occ_text(occ))}</td><td>{e(plan)}</td>")
        for f, t, sc in periods:
            r = lut.get((seg, room, board, basis, plan, occ, sc))
            cls = " class='low'" if r and r.get("confidence") == "low" else ""
            parts.append(f"<td{cls}>{e(cell_text(r, cur))}</td>")
        parts.append("</tr>")
    parts.append("</table>")
    catalog = contract.get("room_catalog") or []
    if catalog:
        parts.append("<h2>Room mapping</h2><table><tr><th>Raw</th><th>Code</th>"
                     "<th>Type</th><th>Category</th><th>View</th><th>Bedding</th>"
                     "<th>Accessible</th><th>Amenities</th><th>Confidence</th></tr>")
        for m in catalog:
            cls = " class='low'" if m.get("confidence") in ("low", "medium") else ""
            parts.append(f"<tr><td>{e(m['raw'])}</td><td>{e(m.get('code') or '')}</td>"
                         f"<td>{e(m.get('room_type') or '')}</td><td>{e(m.get('category') or '')}</td>"
                         f"<td>{e(m.get('view') or '')}</td><td>{e(m.get('bedding') or '')}</td>"
                         f"<td>{'yes' if m.get('accessible') else 'no'}</td>"
                         f"<td>{e(', '.join(m.get('amenities') or []))}</td>"
                         f"<td{cls}>{e(m.get('confidence') or '')}</td></tr>")
        parts.append("</table>")
    parts.append("<h2>Policies</h2>"
                 "<table><tr><th>Type</th><th>Item</th><th>Detail</th><th>Note</th></tr>")
    for line in policy_lines(contract):
        parts.append("<tr>" + "".join(f"<td>{e(str(v or ''))}</td>" for v in line) + "</tr>")
    parts.append("</table>")
    with open(out, "w", encoding="utf-8") as fh:
        fh.write("".join(parts))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("contract")
    ap.add_argument("--format", choices=["xlsx", "html", "csv"], default="xlsx")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    with open(args.contract, encoding="utf-8") as fh:
        contract = json.load(fh)

    day_index = build_day_index(contract)
    if not day_index:
        sys.exit("no season ranges: nothing to render (validate first)")

    {"xlsx": render_xlsx, "html": render_html, "csv": render_csv}[args.format](contract, day_index, args.out)
    print(f"written: {args.out}")


if __name__ == "__main__":
    main()
